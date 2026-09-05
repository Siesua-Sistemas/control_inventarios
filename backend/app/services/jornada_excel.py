"""Generación de los reportes de asistencia/jornada en formato Excel (openpyxl)."""
from __future__ import annotations

from datetime import timedelta, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.recargos import RECARGO_CATEGORIAS, recargo_pct

_HEADER_FILL = PatternFill('solid', fgColor='0E7490')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_ALERT_FILL = PatternFill('solid', fgColor='FEE2E2')
_BOGOTA_OFFSET = timedelta(hours=-5)

_CABECERAS_RECARGOS = [label for _, label, _ in RECARGO_CATEGORIAS]


def _fmt_horas(minutos: int) -> str:
    if not minutos:
        return '0h'
    return f'{minutos // 60}h {minutos % 60:02d}m'


def _hora_local(ts) -> str:
    if ts is None:
        return ''
    if ts.tzinfo is not None:
        ts = ts.astimezone(timezone.utc).replace(tzinfo=None)
    return (ts + _BOGOTA_OFFSET).strftime('%H:%M')


def _novedades_dia(dia: Any) -> str:
    novedades = []
    if any(r.is_manual for r in dia.registros):
        novedades.append('Manual')
    if any(r.ubicacion_no_verificada for r in dia.registros):
        novedades.append('Ubicación no verificada')
    return ' · '.join(novedades)


def _extra_total(recargos_totales: dict[str, int]) -> int:
    return sum(recargos_totales.get(k, 0) for k in (
        'extra_diurna', 'extra_nocturna', 'extra_dominical_diurno', 'extra_dominical_nocturno',
    ))


def _aplicar_cabecera(ws, fila: int, columnas: list[str]) -> None:
    for col, texto in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=col, value=texto)
        celda.font = _HEADER_FONT
        celda.fill = _HEADER_FILL
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _autoancho(ws, anchos: list[int]) -> None:
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _agregar_hoja_referencia(wb: Workbook) -> None:
    ws = wb.create_sheet('Recargos (referencia legal)')
    _aplicar_cabecera(ws, 1, ['Tipo de hora', 'Franja', 'Recargo 2026 (ene-jun)', 'Recargo 2026 (jul-dic)'])
    for i, (clave, label, franja) in enumerate(RECARGO_CATEGORIAS, start=2):
        pct_1 = recargo_pct(clave, 1)
        pct_7 = recargo_pct(clave, 7)
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=franja)
        ws.cell(row=i, column=3, value=f'+{pct_1}%' if pct_1 else '0% (base)')
        ws.cell(row=i, column=4, value=f'+{pct_7}%' if pct_7 else '0% (base)')
    ws.cell(row=len(RECARGO_CATEGORIAS) + 3, column=1, value=(
        'Ley 2466: máx. 2h extra/día y 12h extra/semana '
        '(aprox. proporcional de 15 días para doble turno). '
        'El incumplimiento puede acarrear sanciones de MinTrabajo.'
    )).font = Font(italic=True, size=9, color='64748B')
    _autoancho(ws, [32, 32, 20, 20])


def _to_bytes(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def construir_excel_resumen_mensual(reporte: Any) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen mensual'

    cabeceras = [
        'Empleado', 'Cargo', 'Sede',
        'Días asistidos', 'Días incompletos', 'Días ausentes', 'Total horas',
        *_CABECERAS_RECARGOS,
        'Horas extra totales', 'Días con exceso (>2h extra/día)', 'Periodos con exceso legal',
        'Novedades manuales', 'Novedades ubicación no verificada',
    ]
    _aplicar_cabecera(ws, 1, cabeceras)

    fila = 2
    for e in reporte.empleados:
        valores = [
            f'{e.apellidos} {e.nombres}', e.cargo or '', e.sede or '',
            e.dias_asistidos, e.dias_incompletos, e.dias_ausentes, _fmt_horas(e.total_minutos),
            *[_fmt_horas(e.recargos_totales.get(clave, 0)) for clave, _, _ in RECARGO_CATEGORIAS],
            _fmt_horas(_extra_total(e.recargos_totales)), e.dias_excedidos, e.periodos_excedidos,
            e.novedades_manuales, e.novedades_ubicacion,
        ]
        for col, val in enumerate(valores, start=1):
            ws.cell(row=fila, column=col, value=val)
        if e.dias_excedidos > 0 or e.periodos_excedidos > 0:
            for col in range(1, len(cabeceras) + 1):
                ws.cell(row=fila, column=col).fill = _ALERT_FILL
        fila += 1

    ws.freeze_panes = 'A2'
    _autoancho(ws, [28, 16, 16, 12, 12, 12, 12] + [15] * len(_CABECERAS_RECARGOS) + [14, 14, 14, 14, 16])
    _agregar_hoja_referencia(wb)
    return _to_bytes(wb)


def construir_excel_consolidado_mensual(reporte: Any) -> BytesIO:
    """Una sola hoja con todos los empleados: una fila por día de cada empleado."""
    wb = Workbook()
    ws = wb.active
    ws.title = f'Detalle {reporte.mes}'[:31]

    cabeceras = [
        'Empleado', 'Cargo', 'Sede', 'Fecha', 'Día',
        'Entrada', 'Salida', 'Tiempo neto (sede)',
        *_CABECERAS_RECARGOS, 'Horas extra día', 'Excede 2h/día', 'Novedad',
    ]
    _aplicar_cabecera(ws, 1, cabeceras)

    fila = 2
    for e in reporte.empleados:
        nombre_completo = f'{e.apellidos} {e.nombres}'
        for dia in e.dias:
            entrada = next((r for r in dia.registros if r.tipo == 'entrada'), None)
            salida = next((r for r in dia.registros if r.tipo == 'salida'), None)
            valores = [
                nombre_completo, e.cargo or '', e.sede or '',
                dia.fecha, dia.dia_semana,
                _hora_local(entrada.timestamp if entrada else None),
                _hora_local(salida.timestamp if salida else None),
                dia.tiempo_sede or '',
                *[_fmt_horas(dia.recargos.get(clave, 0)) for clave, _, _ in RECARGO_CATEGORIAS],
                _fmt_horas(dia.extra_min), 'Sí' if dia.excede_diario else '',
                _novedades_dia(dia),
            ]
            for col, val in enumerate(valores, start=1):
                ws.cell(row=fila, column=col, value=val)
            if dia.excede_diario:
                for col in range(1, len(cabeceras) + 1):
                    ws.cell(row=fila, column=col).fill = _ALERT_FILL
            fila += 1

    ws.freeze_panes = 'A2'
    _autoancho(ws, [28, 16, 16, 12, 10, 10, 10, 14] + [15] * len(_CABECERAS_RECARGOS) + [14, 12, 26])
    _agregar_hoja_referencia(wb)
    return _to_bytes(wb)


def construir_excel_detalle_mensual(emp: Any, mes_label: str) -> BytesIO:
    """Reporte individual ('por profesional') de un empleado para el mes."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalle'

    ws.cell(row=1, column=1, value=f'{emp.nombres} {emp.apellidos}').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'{emp.cargo or ""} · {emp.sede or ""} · {mes_label}')
    ws.cell(row=3, column=1, value=(
        f'Días asistidos: {emp.dias_asistidos}   Incompletos: {emp.dias_incompletos}   '
        f'Ausentes: {emp.dias_ausentes}   Total horas: {_fmt_horas(emp.total_minutos)}'
    ))
    fila_inicio = 5
    if emp.dias_excedidos or emp.periodos_excedidos:
        celda = ws.cell(row=4, column=1, value=(
            f'⚠ {emp.dias_excedidos} día(s) superaron el límite de 2h extra/día y '
            f'{emp.periodos_excedidos} periodo(s) superaron el límite legal de horas extra (Ley 2466).'
        ))
        celda.font = Font(color='B91C1C', bold=True)
        fila_inicio = 6

    cabeceras = [
        'Fecha', 'Día', 'Entrada', 'Salida', 'Tiempo neto',
        *_CABECERAS_RECARGOS, 'Horas extra día', 'Excede 2h/día', 'Novedad',
    ]
    _aplicar_cabecera(ws, fila_inicio, cabeceras)

    fila = fila_inicio + 1
    for dia in emp.dias:
        entrada = next((r for r in dia.registros if r.tipo == 'entrada'), None)
        salida = next((r for r in dia.registros if r.tipo == 'salida'), None)
        valores = [
            dia.fecha, dia.dia_semana,
            _hora_local(entrada.timestamp if entrada else None),
            _hora_local(salida.timestamp if salida else None),
            dia.tiempo_sede or '',
            *[_fmt_horas(dia.recargos.get(clave, 0)) for clave, _, _ in RECARGO_CATEGORIAS],
            _fmt_horas(dia.extra_min), 'Sí' if dia.excede_diario else '',
            _novedades_dia(dia),
        ]
        for col, val in enumerate(valores, start=1):
            ws.cell(row=fila, column=col, value=val)
        if dia.excede_diario:
            for col in range(1, len(cabeceras) + 1):
                ws.cell(row=fila, column=col).fill = _ALERT_FILL
        fila += 1

    ws.freeze_panes = f'A{fila_inicio + 1}'
    _autoancho(ws, [12, 10, 10, 10, 14] + [15] * len(_CABECERAS_RECARGOS) + [14, 12, 26])
    _agregar_hoja_referencia(wb)
    return _to_bytes(wb)
